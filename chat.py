import streamlit as st
import pandas as pd
import io
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# Helper function to color DOC column (used for in-app Styler)
def color_doc_column(val):
    """
    Color coding based on DOC ranges:
    0-7: Critical (Dark Red)
    7-15: Low (Orange)
    15-30: Good (Green)
    30-45: Optimal (Yellow/Gold)
    45-60: High (Blue)
    60-90: Excess (Brown)
    >90: Over (Dark Gray)
    Returns CSS style string.
    """
    try:
        if pd.isna(val):
            return ''
        v = float(val)
        if v <= 7:
            return 'background-color: #8B0000; color: white'
        elif v <= 15:
            return 'background-color: #FFA500; color: white'
        elif v <= 30:
            return 'background-color: #228B22; color: white'
        elif v <= 45:
            return 'background-color: #DAA520; color: white'
        elif v <= 60:
            return 'background-color: #4169E1; color: white'
        elif v <= 90:
            return 'background-color: #8B4513; color: white'
        else:
            return 'background-color: #696969; color: white'
    except Exception:
        return ''

# Page configuration
st.set_page_config(
    page_title="Amazon PO Working Report",
    page_icon="📊",
    layout="wide"
)

# Title and description
st.title("📊 Amazon PO Working Report")
st.markdown("Upload your files to process Amazon business reports with inventory and RIS analysis")

# Sidebar for file uploads
st.sidebar.header("📁 Upload Files")

uploaded_business_report = st.sidebar.file_uploader(
    "Business Report CSV",
    type=['csv'],
    help="Upload BusinessReport-11-12-25.csv"
)

uploaded_pm = st.sidebar.file_uploader(
    "Product Master Excel",
    type=['xlsx'],
    help="Upload PM.xlsx"
)

uploaded_inventory = st.sidebar.file_uploader(
    "Inventory CSV",
    type=['csv'],
    help="Upload 400337020433.csv"
)

uploaded_ris = st.sidebar.file_uploader(
    "RIS CSV",
    type=['csv'],
    help="Upload 400334020433.csv"
)

uploaded_fc_state = st.sidebar.file_uploader(
    "FC State Cluster Excel",
    type=['xlsx'],
    help="Upload State FC Cluster.xlsx"
)

# Number of days input with explanation
st.sidebar.markdown("### 📅 Number of Days for Analysis")
st.sidebar.info("""
**What is Number of Days?**

This is the time period covered by your sales data (e.g., 30 or 31 for monthly data).

• **DRR** (Daily Run Rate) = Total Order Items ÷ Days
• **DOC** (Days of Coverage) = Fulfillable Qty ÷ DRR

DOC tells you how many days your current inventory will last at the current sales rate.
""")

no_of_days = st.sidebar.number_input(
    "Enter the number of days",
    min_value=1,
    value=31,
    help="Time period covered by your sales data"
)

# State mapping dictionary (unchanged)
state_mapping = {
    'MAHARASHTRA': 'Maharashtra', 'Maharashtra': 'Maharashtra', 'maharashtra': 'Maharashtra',
    'Maharasthra': 'Maharashtra', 'Thane dt.Maharashtea': 'Maharashtra',
    'GJ': 'Gujarat', 'Gujarat': 'Gujarat', 'GUJARAT': 'Gujarat', 'Gujrat': 'Gujarat', 'Gujurat': 'Gujarat',
    'TELANGANA': 'Telangana', 'Telangana': 'Telangana', 'telangana': 'Telangana', 'TG': 'Telangana',
    'TAMIL NADU': 'Tamil Nadu', 'Tamil Nadu': 'Tamil Nadu', 'TamilNadu': 'Tamil Nadu',
    'Tamilnadu': 'Tamil Nadu', 'TN': 'Tamil Nadu',
    'KARNATAKA': 'Karnataka', 'Karnataka': 'Karnataka', 'karnataka': 'Karnataka',
    'UTTAR PRADESH': 'Uttar Pradesh', 'Uttar Pradesh': 'Uttar Pradesh',
    'uttar pradesh': 'Uttar Pradesh', 'UP': 'Uttar Pradesh',
    'HARYANA': 'Haryana', 'Haryana': 'Haryana', 'HR': 'Haryana',
    'DELHI': 'Delhi', 'Delhi': 'Delhi', 'delhi': 'Delhi', 'New delhi': 'Delhi',
    'New Delhi': 'Delhi', 'NEW DELHI': 'Delhi', 'DL': 'Delhi',
    'ASSAM': 'Assam', 'Assam': 'Assam',
    'WEST BENGAL': 'West Bengal', 'West Bengal': 'West Bengal', 'West bengal': 'West Bengal',
    'MADHYA PRADESH': 'Madhya Pradesh', 'madhya pradesh': 'Madhya Pradesh', 'Madhya Pradesh': 'Madhya Pradesh',
    'PUNJAB': 'Punjab', 'Punjab': 'Punjab', 'punjab': 'Punjab',
    'RAJASTHAN': 'Rajasthan', 'Rajasthan': 'Rajasthan',
    'GOA': 'Goa', 'JHARKHAND': 'Jharkhand', 'UTTARAKHAND': 'Uttarakhand', 'Uttarakhand': 'Uttarakhand',
    'BIHAR': 'Bihar', 'MIZORAM': 'Mizoram', 'MEGHALAYA': 'Meghalaya', 'CHHATTISGARH': 'Chhattisgarh',
    'CHANDIGARH': 'Chandigarh', 'Chandigarh': 'Chandigarh', 'JAMMU & KASHMIR': 'Jammu & Kashmir',
    'NAGALAND': 'Nagaland', 'SIKKIM': 'Sikkim', 'TRIPURA': 'Tripura',
    'ARUNACHAL PRADESH': 'Arunachal Pradesh', 'ANDAMAN & NICOBAR ISLANDS': 'Andaman & Nicobar Islands',
    'PUDUCHERRY': 'Puducherry', 'LADAKH': 'Ladakh',
    'ANDHRA PRADESH': 'Andhra Pradesh', 'Andhra Pradesh': 'Andhra Pradesh', 'Andhra pradesh': 'Andhra Pradesh',
    'ODISHA': 'Odisha', 'Orissa': 'Odisha',
    'DAMAN & DIU': 'Maharashtra', 'DAMAN AND DIU': 'Maharashtra', 'DADRA AND NAGAR HAVELI': 'Maharashtra',
}

# ---------- Create Excel with DOC formatting ----------
def create_excel_with_doc_formatting(df, doc_col_name='DOC'):
    """
    Return bytes of an Excel file where the DOC column cells are filled
    using the color ranges defined in color_doc_column.
    """
    output = io.BytesIO()
    # Use ExcelWriter with openpyxl engine
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Write df to Excel
        df.to_excel(writer, sheet_name="BusinessReport", index=False)
        workbook = writer.book
        worksheet = writer.sheets["BusinessReport"]

        # If DOC column missing, context manager will handle saving; just return bytes after exit
        if doc_col_name not in df.columns:
            pass  # nothing to color

        else:
            # Determine DOC column letter and row range
            doc_col_idx = list(df.columns).index(doc_col_name) + 1  # 1-based
            doc_col_letter = get_column_letter(doc_col_idx)
            first_data_row = 2
            last_data_row = len(df) + 1

            # Define cell fills
            fills = {
                'critical': PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid"),
                'low':      PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
                'good':     PatternFill(start_color="228B22", end_color="228B22", fill_type="solid"),
                'optimal':  PatternFill(start_color="DAA520", end_color="DAA520", fill_type="solid"),
                'high':     PatternFill(start_color="4169E1", end_color="4169E1", fill_type="solid"),
                'excess':   PatternFill(start_color="8B4513", end_color="8B4513", fill_type="solid"),
                'over':     PatternFill(start_color="696969", end_color="696969", fill_type="solid"),
            }

            # Apply fill iteratively (deterministic and clear)
            for row in range(first_data_row, last_data_row + 1):
                cell_ref = f"{doc_col_letter}{row}"
                cell = worksheet[cell_ref]
                try:
                    v = float(cell.value) if cell.value not in (None, '') else None
                except Exception:
                    v = None

                if v is None:
                    continue
                if v <= 7:
                    cell.fill = fills['critical']
                elif v <= 15:
                    cell.fill = fills['low']
                elif v <= 30:
                    cell.fill = fills['good']
                elif v <= 45:
                    cell.fill = fills['optimal']
                elif v <= 60:
                    cell.fill = fills['high']
                elif v <= 90:
                    cell.fill = fills['excess']
                else:
                    cell.fill = fills['over']

        # Freeze header row
        worksheet.freeze_panes = worksheet['A2']
        # Don't call writer.save() - context manager saves on exit

    # context manager closed, buffer contains the workbook
    output.seek(0)
    return output.getvalue()

# ---------- Main processing ----------

# Process button
if st.sidebar.button("🚀 Process Reports", type="primary"):
    if all([uploaded_business_report, uploaded_pm, uploaded_inventory, uploaded_ris, uploaded_fc_state]):
        try:
            with st.spinner("Processing files..."):
                # Load files
                Business_Report = pd.read_csv(uploaded_business_report)
                PM = pd.read_excel(uploaded_pm)
                Inventory = pd.read_csv(uploaded_inventory)
                RIS = pd.read_csv(uploaded_ris)
                FC_State = pd.read_excel(uploaded_fc_state, sheet_name="Sheet2")

                # Process PM data
                PM = PM.iloc[:, 0:7].copy()
                PM.columns = ["ASIN", "B", "C", "Vendor SKU Codes", "Brand Manager", "F", "Brand"]

                # Add Brand Manager and Brand to Business Report
                Business_Report = Business_Report.merge(
                    PM[["ASIN", "Brand Manager"]], how="left",
                    left_on="(Parent) ASIN", right_on="ASIN"
                )
                if "Title" in Business_Report.columns:
                    insert_pos = Business_Report.columns.get_loc("Title")
                    col = Business_Report.pop("Brand Manager")
                    Business_Report.insert(insert_pos, "Brand Manager", col)

                Business_Report = Business_Report.merge(
                    PM[["ASIN", "Brand"]], how="left",
                    left_on="(Parent) ASIN", right_on="ASIN"
                )
                if "Title" in Business_Report.columns:
                    insert_pos = Business_Report.columns.get_loc("Title")
                    col = Business_Report.pop("Brand")
                    Business_Report.insert(insert_pos, "Brand", col)

                # Add Vendor SKU Codes (keep as string-safe)
                Business_Report = Business_Report.merge(
                    PM[["ASIN", "Vendor SKU Codes"]], how="left",
                    left_on="(Parent) ASIN", right_on="ASIN"
                )

                # Process Inventory data
                Inventory.columns = Inventory.columns.str.strip()
                # defensive: ensure there are enough columns
                if len(Inventory.columns) >= 13:
                    lookup_col = Inventory.columns[2]
                    return_col = Inventory.columns[10]
                    return_col_13 = Inventory.columns[12]

                    mi_map = Inventory.set_index(lookup_col)[return_col].to_dict()
                    Business_Report["Current Stock"] = Business_Report["(Parent) ASIN"].map(mi_map)

                    mi_res_map = Inventory.set_index(lookup_col)[return_col_13].to_dict()
                    Business_Report["Reserve Stock"] = Business_Report["(Parent) ASIN"].map(mi_res_map)
                else:
                    # Fallback: create columns if inventory shape unexpected
                    Business_Report["Current Stock"] = np.nan
                    Business_Report["Reserve Stock"] = np.nan

                # Calculate DRR (clean Total Order Items)
                if "Total Order Items" in Business_Report.columns:
                    Business_Report["Total Order Items"] = (
                        Business_Report["Total Order Items"]
                        .astype(str).str.replace("\u00A0", "", regex=False)
                        .str.replace(",", "", regex=False)
                        .str.replace(r"[^\d\.\-]", "", regex=True)
                    )
                    Business_Report["Total Order Items"] = pd.to_numeric(
                        Business_Report["Total Order Items"], errors="coerce"
                    )
                    Business_Report["DRR"] = (Business_Report["Total Order Items"] / no_of_days).round(2)
                else:
                    Business_Report["DRR"] = np.nan

                # Calculate DOC
                Business_Report["Current Stock"] = pd.to_numeric(Business_Report.get("Current Stock"), errors="coerce")
                Business_Report["DRR"] = pd.to_numeric(Business_Report.get("DRR"), errors="coerce")
                Business_Report["DOC"] = (Business_Report["Current Stock"] / Business_Report["DRR"]).round(2)

                # Process RIS data: map FC -> state/cluster
                FC_State.columns = FC_State.columns.str.strip()
                if len(FC_State.columns) >= 4:
                    fc_map = FC_State.set_index(FC_State.columns[0])[FC_State.columns[1]].to_dict()
                    RIS['State'] = RIS['FC'].map(fc_map)

                    fc_map_3 = FC_State.set_index(FC_State.columns[0])[FC_State.columns[2]].to_dict()
                    RIS['Cluster'] = RIS['FC'].map(fc_map_3)

                    fc_map_4 = FC_State.set_index(FC_State.columns[0])[FC_State.columns[3]].to_dict()
                    RIS['StateCluster'] = RIS['FC'].map(fc_map_4)
                else:
                    RIS['State'] = np.nan
                    RIS['Cluster'] = np.nan
                    RIS['StateCluster'] = np.nan

                # Normalize shipping state in RIS
                RIS['Shipping State'] = (
                    RIS['Shipping State'].map(state_mapping)
                    .fillna(RIS['Shipping State'].str.title())
                )

                # RIS Status
                RIS["RIS Status"] = RIS.apply(
                    lambda row: "RIS" if str(row.get("Shipping State", "")).strip().replace(" ", "") ==
                                          str(row.get("State", "")).strip().replace(" ", "")
                                else "Non RIS",
                    axis=1
                )

                # Create pivot table: pivot and avoid multi-index drop pitfalls
                pivot_ris = pd.pivot_table(
                    RIS, values='Shipped Quantity',
                    index=['Merchant SKU', 'Cluster', 'StateCluster'],
                    columns='RIS Status', aggfunc='sum', fill_value=0,
                    margins=True, margins_name='Grand Total'
                )
                pivot_ris = pivot_ris.reset_index()  # convert index into columns
                # If margins row present, it's a row with Merchant SKU == 'Grand Total'
                if 'Grand Total' in pivot_ris['Merchant SKU'].values:
                    grand_total_row = pivot_ris[pivot_ris['Merchant SKU'] == 'Grand Total'].copy()
                    pivot_no_total = pivot_ris[pivot_ris['Merchant SKU'] != 'Grand Total'].copy()
                    # Sort by Grand Total column if present
                    if 'Grand Total' in pivot_no_total.columns:
                        pivot_no_total = pivot_no_total.sort_values(by='Grand Total', ascending=False)
                    pivot_sorted = pd.concat([pivot_no_total, grand_total_row], ignore_index=True)
                else:
                    pivot_sorted = pivot_ris.copy()
                    if 'Grand Total' in pivot_sorted.columns:
                        pivot_sorted = pivot_sorted.sort_values(by='Grand Total', ascending=False)

                pivot_sorted.reset_index(drop=True, inplace=True)

                # Map RIS pivot data to Business Report
                pivot_sorted.columns = pivot_sorted.columns.str.strip()
                if 'Merchant SKU' in pivot_sorted.columns:
                    pivot_sorted_map_cluster = pivot_sorted.set_index('Merchant SKU')['Cluster'].to_dict()
                    Business_Report['RIS High Cluster'] = Business_Report['SKU'].map(pivot_sorted_map_cluster)

                    if 'Grand Total' in pivot_sorted.columns:
                        pivot_sorted_map_qty = pivot_sorted.set_index('Merchant SKU')['Grand Total'].to_dict()
                        Business_Report['RIS QTY'] = Business_Report['SKU'].map(pivot_sorted_map_qty)
                    else:
                        Business_Report['RIS QTY'] = np.nan

                    if 'StateCluster' in pivot_sorted.columns:
                        pivot_sorted_map_state = pivot_sorted.set_index('Merchant SKU')['StateCluster'].to_dict()
                        Business_Report['RIS State'] = Business_Report['SKU'].map(pivot_sorted_map_state)
                    else:
                        Business_Report['RIS State'] = np.nan
                else:
                    Business_Report['RIS High Cluster'] = np.nan
                    Business_Report['RIS QTY'] = np.nan
                    Business_Report['RIS State'] = np.nan

                # PO State
                Business_Report['PO State'] = Business_Report['DOC'].apply(
                    lambda x: "Create A PO" if pd.notna(x) and x <= 30 else "We Have Stock"
                )

                # Case Pack handling - use np.where to avoid replace downcasting warning
                PM1 = pd.read_excel(uploaded_pm)
                PM1.columns = PM1.columns.str.strip()
                if PM1.shape[1] > 11:
                    casepack_map = PM1.set_index(PM1.columns[0])[PM1.columns[11]].to_dict()
                    Business_Report['Case Pack'] = Business_Report['(Parent) ASIN'].map(casepack_map)
                else:
                    Business_Report['Case Pack'] = 1

                # Replace "NO CASE PACK" with 1 using np.where (avoids future downcast warning) and coerce to numeric
                Business_Report['Case Pack'] = np.where(Business_Report['Case Pack'] == "NO CASE PACK", 1, Business_Report['Case Pack'])
                Business_Report['Case Pack'] = pd.to_numeric(Business_Report['Case Pack'], errors='coerce').fillna(1).astype(float)

                # ---------- Ensure object columns are string to avoid pyarrow serialization errors ----------
                # Convert all object dtype columns to string (safe and avoids mixed-type serialization issues)
                obj_cols = Business_Report.select_dtypes(include=['object']).columns.tolist()
                for c in obj_cols:
                    Business_Report[c] = Business_Report[c].astype(str).fillna('')

                # Reformat numeric columns consistently
                Business_Report['DOC'] = pd.to_numeric(Business_Report.get('DOC'), errors='coerce').round(2)
                Business_Report['DRR'] = pd.to_numeric(Business_Report.get('DRR'), errors='coerce').round(2)
                Business_Report['Current Stock'] = pd.to_numeric(Business_Report.get('Current Stock'), errors='coerce')

                # Create Excel bytes with DOC conditional formatting
                excel_bytes = create_excel_with_doc_formatting(Business_Report, doc_col_name='DOC')

            st.success("✅ Processing completed successfully!")

            # Display DOC Color Legend
            st.markdown("### 🎨 DOC Color Legend")
            col1, col2, col3, col4, col5, col6 = st.columns(6)
            with col1:
                st.markdown('<div style="background-color: #8B0000; color: white; padding: 10px; text-align: center; border-radius: 5px;"><b>0-7 days</b><br>(Critical)</div>', unsafe_allow_html=True)
            with col2:
                st.markdown('<div style="background-color: #FFA500; color: white; padding: 10px; text-align: center; border-radius: 5px;"><b>7-15 days</b><br>(Low)</div>', unsafe_allow_html=True)
            with col3:
                st.markdown('<div style="background-color: #228B22; color: white; padding: 10px; text-align: center; border-radius: 5px;"><b>15-30 days</b><br>(Good)</div>', unsafe_allow_html=True)
            with col4:
                st.markdown('<div style="background-color: #DAA520; color: white; padding: 10px; text-align: center; border-radius: 5px;"><b>30-45 days</b><br>(Optimal)</div>', unsafe_allow_html=True)
            with col5:
                st.markdown('<div style="background-color: #4169E1; color: white; padding: 10px; text-align: center; border-radius: 5px;"><b>45-60 days</b><br>(High)</div>', unsafe_allow_html=True)
            with col6:
                st.markdown('<div style="background-color: #8B4513; color: white; padding: 10px; text-align: center; border-radius: 5px;"><b>60-90 days</b><br>(Excess)</div>', unsafe_allow_html=True)

            st.markdown("---")

            # Display results in tabs
            tab1, tab2, tab3 = st.tabs(["📊 Business Report", "📦 RIS Analysis", "📈 Summary Statistics"])

            with tab1:
                st.subheader("Processed Business Report")

                # Use Styler.apply for DOC column styling to avoid deprecated applymap
                if 'DOC' in Business_Report.columns:
                    styled_df = Business_Report.style.apply(
                        lambda col: [color_doc_column(v) for v in col] if col.name == 'DOC' else [''] * len(col),
                        axis=0
                    )
                else:
                    styled_df = Business_Report.style

                # Show in-app table
                st.dataframe(styled_df, use_container_width=True)

                # CSV download
                csv = Business_Report.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="⬇️ Download Business Report CSV",
                    data=csv,
                    file_name="processed_business_report.csv",
                    mime="text/csv"
                )

                # Excel download (with DOC formatting)
                st.download_button(
                    label="⬇️ Download Business Report (Excel with DOC formatting)",
                    data=excel_bytes,
                    file_name="processed_business_report_with_DOC_formatting.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

            with tab2:
                st.subheader("RIS Pivot Analysis")
                st.dataframe(pivot_sorted, use_container_width=True)

                csv_pivot = pivot_sorted.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="⬇️ Download RIS Pivot CSV",
                    data=csv_pivot,
                    file_name="ris_pivot_analysis.csv",
                    mime="text/csv"
                )

            with tab3:
                st.subheader("Summary Statistics")

                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Total Products", len(Business_Report))
                with col2:
                    st.metric("Create PO Items", int(len(Business_Report[Business_Report['PO State'] == 'Create A PO'])))
                with col3:
                    st.metric("In Stock Items", int(len(Business_Report[Business_Report['PO State'] == 'We Have Stock'])))
                with col4:
                    avg_doc = Business_Report['DOC'].mean()
                    st.metric("Avg DOC", f"{avg_doc:.2f}" if pd.notna(avg_doc) else "N/A")

                # DOC Distribution by Category
                st.subheader("DOC Distribution by Category")
                doc_bins = [0, 7, 15, 30, 45, 60, 90, float('inf')]
                doc_labels = ['0-7 (Critical)', '7-15 (Low)', '15-30 (Good)', '30-45 (Optimal)', '45-60 (High)', '60-90 (Excess)', '90+ (Excess)']
                doc_categories = pd.cut(
                    Business_Report['DOC'].fillna(-1),
                    bins=doc_bins,
                    labels=doc_labels,
                    include_lowest=True
                )
                doc_dist = doc_categories.value_counts().reindex(doc_labels).fillna(0).astype(int)
                st.bar_chart(doc_dist)

                st.subheader("PO State Distribution")
                po_dist = Business_Report['PO State'].value_counts()
                st.bar_chart(po_dist)

                st.subheader("Top 10 Products by DRR")
                if 'DRR' in Business_Report.columns:
                    top_products = Business_Report.nlargest(10, 'DRR')[['Title', 'DRR', 'Current Stock', 'DOC']].copy()
                    # style DOC column
                    if 'DOC' in top_products.columns:
                        styled_top = top_products.style.apply(
                            lambda col: [color_doc_column(v) for v in col] if col.name == 'DOC' else [''] * len(col),
                            axis=0
                        )
                    else:
                        styled_top = top_products.style
                    st.dataframe(styled_top, use_container_width=True)

                # Critical Items (DOC <= 7)
                critical_items = Business_Report[pd.to_numeric(Business_Report['DOC'], errors='coerce') <= 7]
                if len(critical_items) > 0:
                    st.subheader(f"⚠️ Critical Items (DOC ≤ 7 days) - {len(critical_items)} items")
                    critical_display = critical_items[['SKU', 'Title', 'Brand', 'Current Stock', 'DRR', 'DOC', 'PO State']].head(100)
                    if 'DOC' in critical_display.columns:
                        styled_critical = critical_display.style.apply(
                            lambda col: [color_doc_column(v) for v in col] if col.name == 'DOC' else [''] * len(col),
                            axis=0
                        )
                    else:
                        styled_critical = critical_display.style
                    st.dataframe(styled_critical, use_container_width=True)

        except Exception as e:
            st.error(f"❌ An error occurred: {str(e)}")
            st.exception(e)
    else:
        st.warning("⚠️ Please upload all required files before processing")

else:
    st.info("👈 Upload all required files in the sidebar and click 'Process Reports' to begin")

    # Instructions
    st.markdown("### 📋 Instructions")
    st.markdown("""
    1. Upload the following files in the sidebar:
        - **Business Report CSV**: Your main business report
        - **Product Master Excel**: Product master data (PM.xlsx)
        - **Inventory CSV**: Inventory data
        - **RIS CSV**: RIS order data
        - **FC State Cluster Excel**: Fulfillment center and state mapping
    2. Set the **Number of Days** for DRR calculation (default: 30)
    3. Click **Process Reports** to generate the analysis
    4. View results in different tabs and download processed files

    ### 📊 Output Features
    - Brand Manager and Brand mapping
    - Current and Reserve Stock levels
    - Daily Run Rate (DRR) calculation
    - Days of Coverage (DOC) calculation
    - RIS (Regional In-State) analysis
    - PO recommendations
    - Case Pack information
    """)

